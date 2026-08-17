#!/usr/bin/env python3
"""HUD measurement software socket client.

Switch one or more configuration files and execute measurement commands.
"""

from __future__ import annotations

import argparse
import ast
import re
import socket
import sys
import time
from dataclasses import dataclass
from pathlib import Path


TERMINATOR = b"%"


class HudProtocolError(RuntimeError):
    """Raised when the HUD software returns an unexpected response."""


@dataclass(frozen=True)
class TestResult:
    config: str
    command: str
    response: str


def parse_t24_rows(response: str) -> list[list[float]]:
    """Extract only column 2 from each comma-separated seven-column t24 row."""
    match = re.match(r"^t24_Result(?:\(\d+\))?:(.*?)%?$", response.strip(), re.DOTALL)
    if not match:
        raise HudProtocolError(f"Invalid t24 response format: {response!r}")

    groups = [group.strip() for group in match.group(1).split(",") if group.strip()]
    if not groups:
        raise HudProtocolError("t24 returned no data rows")

    rows: list[list[float]] = []
    for index, group in enumerate(groups, start=1):
        parts = group.split()
        if len(parts) != 7:
            raise HudProtocolError(
                f"t24 row {index} has {len(parts)} columns; expected 7: {group!r}"
            )
        try:
            rows.append([float(parts[1])])
        except ValueError as exc:
            raise HudProtocolError(
                f"t24 row {index} has a non-numeric second-column value: {parts[1]!r}"
            ) from exc
    return rows


def parse_t6_rows(response: str) -> list[list[float]]:
    """Extract only column 2 from each comma/newline-separated t6 data row."""
    match = re.match(r"^t6_Result(?:\(\d+\))?:(.*?)%?$", response.strip(), re.DOTALL)
    if not match:
        raise HudProtocolError(f"Invalid t6 response format: {response!r}")

    groups = [
        group.strip()
        for group in re.split(r",|\r?\n", match.group(1))
        if group.strip()
    ]
    if not groups:
        raise HudProtocolError("t6 returned no data rows")

    rows: list[list[float]] = []
    for index, group in enumerate(groups, start=1):
        parts = group.split()
        if len(parts) < 2:
            raise HudProtocolError(f"t6 row {index} has no second column: {group!r}")
        try:
            rows.append([float(parts[1])])
        except ValueError as exc:
            raise HudProtocolError(
                f"t6 row {index} has a non-numeric second-column value: {parts[1]!r}"
            ) from exc
    return rows


def export_results_excel(results: list[TestResult], output_path: str) -> Path:
    """Export t24/t6 column-2 values into separate measurement sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise HudProtocolError(
            "Excel export requires openpyxl; install it with: python3 -m pip install openpyxl"
        ) from exc

    supported = {"t24": parse_t24_rows, "t6": parse_t6_rows}
    grouped: dict[str, list[tuple[str, list[list[float]]]]] = {}
    indexes: dict[str, dict[str, list[list[float]]]] = {}
    for result in results:
        command = result.command.split("/", 1)[0]
        parser = supported.get(command)
        if parser is None:
            continue
        if command not in grouped:
            grouped[command] = []
            indexes[command] = {}
        if result.config not in indexes[command]:
            rows: list[list[float]] = []
            indexes[command][result.config] = rows
            grouped[command].append((result.config, rows))
        indexes[command][result.config].extend(parser(result.response))

    if not grouped:
        raise HudProtocolError("No t24 or t6 results were available for Excel export")

    output = Path(output_path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    fills = ["E2F0D9", "DDEBF7", "FFF2CC", "FCE4D6"]
    header_border = Border(
        left=Side(style="thin", color="7F8C8D"),
        right=Side(style="thin", color="7F8C8D"),
        top=Side(style="thin", color="7F8C8D"),
        bottom=Side(style="thin", color="7F8C8D"),
    )
    data_border = Border(
        left=Side(style="thin", color="D9E1F2"),
        right=Side(style="thin", color="D9E1F2"),
        top=Side(style="thin", color="D9E1F2"),
        bottom=Side(style="thin", color="D9E1F2"),
    )

    for command, blocks in grouped.items():
        sheet = workbook.create_sheet(f"{command}测试结果")
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        for block_index, (config, rows) in enumerate(blocks):
            column = 1 + block_index * 2
            header = sheet.cell(row=1, column=column, value=f"{config}（配置文件名称）")
            header.fill = PatternFill("solid", fgColor=fills[block_index % len(fills)])
            header.font = Font(bold=True, color="1F2937")
            header.alignment = Alignment(horizontal="center", vertical="center")
            header.border = header_border
            sheet.column_dimensions[header.column_letter].width = 24

            for row_index, row in enumerate(rows, start=2):
                cell = sheet.cell(row=row_index, column=column, value=row[0])
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = data_border
                cell.number_format = "0.###############"
        sheet.row_dimensions[1].height = 24

    workbook.save(output)
    return output


def export_t24_excel(results: list[TestResult], output_path: str) -> Path:
    """Backward-compatible wrapper for callers that export t24 results."""
    return export_results_excel(results, output_path)


class HudClient:
    def __init__(
        self,
        host: str = "127.0.0.1",
        port: int = 5555,
        timeout: float = 60.0,
        response_idle: float = 1.0,
        debug: bool = False,
    ):
        self.host = host
        self.port = port
        self.timeout = timeout
        self.response_idle = response_idle
        self.debug = debug
        self._socket: socket.socket | None = None
        self._buffer = bytearray()
        self._peer_closed = False

    def connect(self) -> None:
        if self._socket is not None and not self._peer_closed:
            return
        if self._socket is not None:
            self._socket.close()
        self._socket = socket.create_connection((self.host, self.port), self.timeout)
        self._socket.settimeout(self.timeout)
        self._peer_closed = False

    def close(self) -> None:
        if self._socket is not None:
            self._socket.close()
            self._socket = None
        self._buffer.clear()
        self._peer_closed = False

    def __enter__(self) -> "HudClient":
        self.connect()
        return self

    def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
        self.close()

    def request(self, command: str) -> str:
        """Send one command and receive through the first '%' terminator."""
        if self._socket is None:
            raise RuntimeError("HUD client is not connected")
        if self._peer_closed:
            self.connect()

        wire_command = command.strip()
        if not wire_command:
            raise ValueError("Command cannot be empty")

        if self.debug:
            print(f"TX {wire_command!r}", flush=True)
        self._socket.sendall(wire_command.encode("utf-8"))
        response = self._receive_message()
        if self.debug:
            print(f"RX {response!r}", flush=True)
        return response

    def _receive_message(self) -> str:
        if self._socket is None:
            raise RuntimeError("HUD client is not connected")

        received_any = bool(self._buffer)
        try:
            while True:
                end = self._buffer.find(TERMINATOR)
                if end >= 0:
                    message = bytes(self._buffer[: end + 1])
                    del self._buffer[: end + 1]
                    return message.decode("utf-8", errors="replace").strip()

                try:
                    chunk = self._socket.recv(4096)
                except socket.timeout:
                    if received_any:
                        message = bytes(self._buffer)
                        self._buffer.clear()
                        return message.decode("utf-8", errors="replace").strip()
                    raise

                if not chunk:
                    if received_any:
                        message = bytes(self._buffer)
                        self._buffer.clear()
                        self._peer_closed = True
                        return message.decode("utf-8", errors="replace").strip()
                    self._peer_closed = True
                    raise ConnectionError("HUD software closed the connection before returning data")

                self._buffer.extend(chunk)
                received_any = True
                self._socket.settimeout(self.response_idle)
        finally:
            if self._socket is not None and not self._peer_closed:
                self._socket.settimeout(self.timeout)

    def switch_config(self, config_name: str) -> None:
        name = config_name.strip()
        if not name:
            raise ValueError("Configuration name cannot be empty")
        if name.lower().endswith(".ini"):
            name = name[:-4]

        response = self.request(f"c-{name}%")
        if response.rstrip("%") != "OK":
            raise HudProtocolError(f"Failed to switch configuration {name!r}: {response}")

    def measure(self, command: str) -> str:
        response = self.request(command)
        expected_prefix = command.split("/", 1)[0] + "_Result"
        if response in {"Error0%", "Error1%", "Fail%"}:
            raise HudProtocolError(f"Measurement {command!r} failed: {response}")
        if not response.startswith(expected_prefix):
            raise HudProtocolError(
                f"Unexpected response to {command!r}; expected {expected_prefix!r}, got {response!r}"
            )
        return response


def run_tests(
    client: HudClient,
    configs: list[str],
    commands: list[str],
    switch_delay: float,
) -> list[TestResult]:
    results: list[TestResult] = []
    for config in configs:
        print(f"[{config}] switching configuration...", flush=True)
        client.switch_config(config)
        if switch_delay:
            time.sleep(switch_delay)

        for command in commands:
            print(f"[{config}] running {command}...", flush=True)
            response = client.measure(command)
            results.append(TestResult(config, command, response))
            print(f"[{config}] {response}", flush=True)
    return results


def run_cases(
    client: HudClient,
    cases: list[tuple[str, str]],
    switch_delay: float,
) -> list[TestResult]:
    """Run explicitly paired configuration/command cases."""
    results: list[TestResult] = []
    for config, command in cases:
        print(f"[{config}] switching configuration...", flush=True)
        client.switch_config(config)
        if switch_delay:
            time.sleep(switch_delay)
        print(f"[{config}] running {command}...", flush=True)
        response = client.measure(command)
        results.append(TestResult(config, command, response))
        print(f"[{config}] {response}", flush=True)
    return results


def load_test_plan(path: str) -> list[tuple[str, list[str]]]:
    """Read a literal TEST_PLAN assignment without executing the config file."""
    config_path = Path(path)
    try:
        tree = ast.parse(config_path.read_text(encoding="utf-8"), filename=str(config_path))
    except (OSError, SyntaxError) as exc:
        raise ValueError(f"Cannot read test plan {config_path}: {exc}") from exc

    raw_plan = None
    for node in tree.body:
        if isinstance(node, ast.Assign) and any(
            isinstance(target, ast.Name) and target.id == "TEST_PLAN" for target in node.targets
        ):
            try:
                raw_plan = ast.literal_eval(node.value)
            except (ValueError, TypeError, SyntaxError) as exc:
                raise ValueError("TEST_PLAN must contain literal strings, lists, and tuples only") from exc
            break

    if raw_plan is None:
        raise ValueError(f"TEST_PLAN was not found in {config_path}")
    if not isinstance(raw_plan, (list, tuple)) or not raw_plan:
        raise ValueError("TEST_PLAN must be a non-empty list")

    plan: list[tuple[str, list[str]]] = []
    for index, item in enumerate(raw_plan, start=1):
        if not isinstance(item, (list, tuple)) or len(item) != 2:
            raise ValueError(f"TEST_PLAN item {index} must be (config, [commands])")
        config, commands = item
        if not isinstance(config, str) or not config.strip():
            raise ValueError(f"TEST_PLAN item {index} has an invalid config name")
        if not isinstance(commands, (list, tuple)) or not commands:
            raise ValueError(f"TEST_PLAN item {index} must contain at least one command")
        if not all(isinstance(command, str) and command.strip() for command in commands):
            raise ValueError(f"TEST_PLAN item {index} contains an invalid command")
        plan.append((config.strip(), [command.strip() for command in commands]))
    return plan


def run_test_plan(
    client: HudClient,
    plan: list[tuple[str, list[str]]],
    switch_delay: float,
) -> list[TestResult]:
    results: list[TestResult] = []
    for config, commands in plan:
        print(f"[{config}] switching configuration...", flush=True)
        client.switch_config(config)
        if switch_delay:
            time.sleep(switch_delay)
        for command in commands:
            print(f"[{config}] running {command}...", flush=True)
            response = client.measure(command)
            results.append(TestResult(config, command, response))
            print(f"[{config}] {response}", flush=True)
    return results


def parse_case(value: str) -> tuple[str, str]:
    try:
        config, command = value.rsplit(":", 1)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(
            "Case must use CONFIG:COMMAND format, for example config1:t1"
        ) from exc
    if not config.strip() or not command.strip():
        raise argparse.ArgumentTypeError("Configuration and command cannot be empty")
    return config.strip(), command.strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Switch HUD software configurations and run measurement commands."
    )
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=5555)
    parser.add_argument("--timeout", type=float, default=60.0)
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Print exact transmitted commands and raw received responses.",
    )
    parser.add_argument(
        "--config",
        action="append",
        help="Configuration filename without extension; repeat for multiple configurations.",
    )
    parser.add_argument(
        "--test",
        action="append",
        help="Measurement command such as t1, t3, or t11/10/20; repeat as needed.",
    )
    parser.add_argument(
        "--case",
        action="append",
        type=parse_case,
        help="Paired CONFIG:COMMAND case; repeat as needed, e.g. --case config1:t1.",
    )
    parser.add_argument(
        "--plan",
        help="Python-style test plan file containing a literal TEST_PLAN list.",
    )
    parser.add_argument(
        "--switch-delay",
        type=float,
        default=0.2,
        help="Seconds to wait after switching configuration (default: 0.2).",
    )
    parser.add_argument(
        "--output",
        default="hud_t24_results.xlsx",
        help="Excel output path for t24 results (default: hud_t24_results.xlsx).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.plan and not args.case and not (args.config and args.test):
        print(
            "ERROR: use --plan FILE, --case CONFIG:COMMAND, or both --config and --test",
            file=sys.stderr,
        )
        return 2
    try:
        with HudClient(args.host, args.port, args.timeout, debug=args.debug) as client:
            if args.plan:
                results = run_test_plan(client, load_test_plan(args.plan), args.switch_delay)
            elif args.case:
                results = run_cases(client, args.case, args.switch_delay)
            else:
                results = run_tests(client, args.config, args.test, args.switch_delay)
        if any(result.command.split("/", 1)[0] in {"t24", "t6"} for result in results):
            output = export_results_excel(results, args.output)
            print(f"Excel saved: {output}", flush=True)
    except (OSError, HudProtocolError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
